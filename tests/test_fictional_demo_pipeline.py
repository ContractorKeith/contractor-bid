from __future__ import annotations

import shutil
import tempfile
import unittest
import zipfile
from contextlib import redirect_stderr, redirect_stdout
from io import StringIO
from pathlib import Path

from openpyxl import load_workbook

from contractor_bid.cli import main
from contractor_bid.util import read_json


REPO_ROOT = Path(__file__).resolve().parents[1]
SAMPLE = REPO_ROOT / "examples" / "fictional-fences-gates-demo"


class FictionalFencesGatesDemoPipelineTest(unittest.TestCase):
    def run_cli(self, args: list[str]) -> str:
        stdout = StringIO()
        stderr = StringIO()
        with redirect_stdout(stdout), redirect_stderr(stderr):
            code = main(args)
        self.assertEqual(
            code,
            0,
            "contractor-bid "
            f"{' '.join(args)} failed\nSTDOUT:\n{stdout.getvalue()}\n"
            f"STDERR:\n{stderr.getvalue()}",
        )
        return stdout.getvalue()

    def test_fictional_fences_gates_demo_runs_full_pipeline(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            project = root / "bids" / "070126-fictional-cedar-park-fence"
            project_name = "Fictional Cedar Park Service Yard"

            self.run_cli(
                [
                    "--root",
                    str(root),
                    "new",
                    str(project),
                    "--profile",
                    "fences-gates",
                    "--project-name",
                    project_name,
                    "--bid-due",
                    "2026-07-01 14:00",
                    "--gc",
                    "Demo GC Estimating <estimating@example.invalid>",
                    "--address",
                    "100 Demo Service Road, Example City, ST",
                ]
            )
            shutil.copytree(SAMPLE, project, dirs_exist_ok=True)

            self.run_cli(["--root", str(root), "triage", str(project), "--profile", "fences-gates"])
            work = project / "bid-package-working"
            hits = read_json(work / "text-extracts" / "page-hits.json")["hits"]
            statuses = {(hit["source_file"], hit["pdf_page"]): hit["status"] for hit in hits}
            self.assertEqual(
                statuses[("fictional-cedar-park-bid-set.pdf", 1)],
                "primary-review",
            )
            self.assertEqual(
                statuses[("fictional-cedar-park-bid-set.pdf", 4)],
                "exclude-review",
            )
            self.assertEqual(
                statuses[("fictional-cedar-park-bid-set.pdf", 5)],
                "flag-review",
            )

            candidate_pages = (work / "takeoff" / "candidate-pages.md").read_text(
                encoding="utf-8"
            )
            self.assertIn("chain link", candidate_pages)
            self.assertIn("exclude-review", candidate_pages)
            self.assertIn("flag-review", candidate_pages)

            sources = read_json(work / "takeoff" / "scope-pages-sources.json")
            self.assertEqual(len(sources["scope_pages"]), 3)
            self.assertEqual(len(sources["spec_pages"]), 1)
            excluded_reasons = " ".join(
                item["reason"] for item in sources["excluded_or_reference_only"]
            ).lower()
            for term in ("bollards", "temporary fence", "silt fence", "tree protection fence"):
                self.assertIn(term, excluded_reasons)
            self.assertIn("guardrail and handrail", excluded_reasons)

            self.run_cli(["--root", str(root), "build-packets", str(project)])
            for rel in (
                "00-Bid-Scope-Summary.md",
                "scope-pages.pdf",
                "scope-pages-index.md",
                "spec-pages.pdf",
                "spec-pages-index.md",
                "scope-and-spec-pages.pdf",
            ):
                self.assertTrue((work / rel).exists(), rel)

            summary = (work / "00-Bid-Scope-Summary.md").read_text(encoding="utf-8")
            self.assertIn("420 LF galvanized chain link fence", summary)
            self.assertIn("Bollards", summary)
            self.assertIn("Guardrail and handrail", summary)

            self.run_cli(
                [
                    "--root",
                    str(root),
                    "build-workbook",
                    str(project),
                    "--profile",
                    "fences-gates",
                ]
            )
            workbook_path = work / "01-Takeoff-Worksheet-REV1.xlsx"
            self.assertTrue(workbook_path.exists())

            takeoff = read_json(work / "takeoff" / "fictional-cedar-park-service-yard.json")
            included_text = " ".join(
                " ".join(str(row.get(key, "")) for key in ("section", "item", "description"))
                for row in takeoff["bom"]
                if str(row.get("status", "")).startswith("Included")
            ).lower()
            for term in (
                "bollard",
                "temporary fence",
                "silt fence",
                "tree protection fence",
                "concrete",
                "paving",
                "handrail",
                "guardrail",
            ):
                self.assertNotIn(term, included_text)

            workbook = load_workbook(workbook_path)
            bom_values = [
                [cell.value for cell in row]
                for row in workbook["BOM"].iter_rows(min_row=5, max_col=6)
                if row[1].value
            ]
            status_by_item = {str(row[1]): str(row[5]) for row in bom_values}
            self.assertEqual(
                status_by_item["Bollards, concrete, and paving restoration"],
                "Excluded",
            )
            self.assertEqual(
                status_by_item["Guardrail and handrail"],
                "Flag / Review",
            )
            self.assertEqual(
                status_by_item["Gate operator and access-control interface"],
                "Alternate / Not Included",
            )

            self.run_cli(
                [
                    "--root",
                    str(root),
                    "check",
                    str(project),
                    "--profile",
                    "fences-gates",
                    "--today",
                    "2026-06-30",
                ]
            )
            alerts = (work / "ALERTS.md").read_text(encoding="utf-8")
            self.assertIn("URGENT: bid due 2026-07-01 - 1 day(s) left.", alerts)
            self.assertIn("Addendum/revision-named files found", alerts)
            self.assertIn("bollard", alerts.lower())
            self.assertIn("guardrail", alerts.lower())

            self.run_cli(["--root", str(root), "package-sendoff", str(project)])
            package_name = "070126-fictional-cedar-park-fence-supplier-sendoff"
            zip_path = work / "supplier-sendoff" / f"{package_name}.zip"
            self.assertTrue(zip_path.exists())
            expected_zip_names = {
                f"{package_name}/00-Bid-Scope-Summary.md",
                f"{package_name}/00-Scope-Reference-Index.md",
                f"{package_name}/01-Takeoff-Worksheet-REV1.xlsx",
                f"{package_name}/02 - Proposal Letter.md",
                f"{package_name}/ALERTS.md",
                f"{package_name}/README.md",
                f"{package_name}/scope-and-spec-pages.pdf",
                f"{package_name}/scope-pages-index.md",
                f"{package_name}/scope-pages.pdf",
                f"{package_name}/spec-pages-index.md",
                f"{package_name}/spec-pages.pdf",
                f"{package_name}/takeoff/review-pages.md",
                f"{package_name}/takeoff/triage-scope-signals.md",
            }
            with zipfile.ZipFile(zip_path) as zf:
                self.assertEqual(set(zf.namelist()), expected_zip_names)
                self.assertFalse(any("/bid-docs/" in name for name in zf.namelist()))
                self.assertFalse(any("scope-pages-sources" in name for name in zf.namelist()))
                self.assertFalse(any("candidate-pages" in name for name in zf.namelist()))

            leak_markers = {
                str(root),
                str(root.resolve()),
                str(project),
                str(project.resolve()),
            }
            for markdown in work.rglob("*.md"):
                text = markdown.read_text(encoding="utf-8")
                for marker in leak_markers:
                    self.assertNotIn(marker, text, markdown)


if __name__ == "__main__":
    unittest.main()
