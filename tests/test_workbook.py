from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from contractor_bid.util import write_json
from contractor_bid.workbook import build_workbook


class WorkbookTest(unittest.TestCase):
    def test_gate_specs_fallback_is_not_used(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            project = Path(tmp) / "070126-sample"
            takeoff = project / "bid-package-working" / "takeoff"
            takeoff.mkdir(parents=True)
            write_json(
                takeoff / "sample.json",
                {
                    "project_name": "Sample",
                    "bom": [],
                    "gate_specs": [
                        {
                            "category": "Old fence fallback",
                            "requirement": "Should not appear",
                            "source": "legacy",
                            "notes": "",
                        }
                    ],
                },
            )

            out = build_workbook(project)
            workbook = load_workbook(out)
            values = [
                cell.value
                for row in workbook["Scope & Specs"].iter_rows()
                for cell in row
                if cell.value
            ]
            self.assertNotIn("Old fence fallback", values)
            self.assertNotIn("Should not appear", values)

    def test_suggested_sources_json_is_not_treated_as_takeoff_json(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            project = Path(tmp) / "070126-sample"
            takeoff = project / "bid-package-working" / "takeoff"
            takeoff.mkdir(parents=True)
            write_json(
                takeoff / "scope-pages-sources.suggested.json",
                {"scope_pages": [{"source_pdf": "bid-docs/source.pdf", "pdf_page": 1}]},
            )
            write_json(
                takeoff / "sample.json",
                {
                    "project_name": "Sample",
                    "bom": [],
                    "scope_specs": [],
                },
            )

            out = build_workbook(project)
            self.assertTrue(out.exists())

    def test_bom_title_spans_all_columns(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            project = Path(tmp) / "070126-sample"
            takeoff = project / "bid-package-working" / "takeoff"
            takeoff.mkdir(parents=True)
            write_json(
                takeoff / "sample.json",
                {
                    "project_name": "Sample",
                    "bom": [
                        {
                            "section": "32",
                            "item": "Fence",
                            "description": "Chain link fence",
                            "qty": 10,
                            "uom": "LF",
                            "status": "Included",
                        }
                    ],
                    "scope_specs": [],
                },
            )

            out = build_workbook(project)
            workbook = load_workbook(out)
            merged_ranges = {str(rng) for rng in workbook["BOM"].merged_cells.ranges}

            self.assertIn("A1:M1", merged_ranges)
            self.assertIn("A2:M2", merged_ranges)


if __name__ == "__main__":
    unittest.main()
