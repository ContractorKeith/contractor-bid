from __future__ import annotations

from pathlib import Path
from typing import Any

from .util import read_json

# Import openpyxl lazily so the rest of the CLI (doctor, triage, check, the
# tracker, etc.) still works when openpyxl is not installed. The error is raised
# at call time in build_workbook() instead of killing the whole CLI on import.
try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.worksheet import Worksheet

    OPENPYXL_AVAILABLE = True
except ImportError:  # pragma: no cover - exercised only when openpyxl is absent
    OPENPYXL_AVAILABLE = False
    Workbook = Alignment = Font = PatternFill = get_column_letter = Worksheet = None  # type: ignore


NAVY = "1F4E78"
TEAL = "173F4F"
TEAL_LIGHT = "EAF3F6"
GREEN_HDR = "D9EAD3"
GREEN_TXT = "1F4E1F"
INPUT_YELLOW = "FFF2CC"
INCLUDED_FILL = "E2EFDA"
INCLUDED_TXT = "215E21"
EXCLUDED_FILL = "FCE4D6"
EXCLUDED_TXT = "9C3D0A"
FLAG_FILL = "FFF2CC"
FLAG_TXT = "7F6000"

WATCH_LIST = [
    ("Latest addendum", "Reconcile quantities and scope to the latest addendum before issuing a price."),
    ("Scope boundary", "Do not include adjacent scopes unless the profile or user explicitly approves them."),
    ("Manual measurements", "PDF text hits are not takeoff quantities; verify LF/SF/CY/counts from plans."),
    ("Tax and freight", "Verify local tax, freight, lift/access, and project-specific procurement costs."),
    ("Alternates", "Keep base bid and alternates separate in the BOM and proposal letter."),
]


def fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def style_header_row(ws: Worksheet, row: int, ncols: int, start_col: int = 1) -> None:
    for col in range(start_col, start_col + ncols):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill(NAVY)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def section_header(ws: Worksheet, row: int, text: str, ncols: int) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.fill = fill(GREEN_HDR)
    cell.font = Font(bold=True, color=GREEN_TXT)


def title_block(ws: Worksheet, text: str, subtitle: str, ncols: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    cell = ws.cell(row=1, column=1, value=text)
    cell.fill = fill(TEAL)
    cell.font = Font(bold=True, color="FFFFFF", size=16)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if subtitle:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
        sub = ws.cell(row=2, column=1, value=subtitle)
        sub.fill = fill(TEAL_LIGHT)
        sub.font = Font(italic=True, color=TEAL, size=10)
        sub.alignment = Alignment(horizontal="center", vertical="center")


def write_rows(ws: Worksheet, start_row: int, rows: list[list[Any]], wrap: bool = True) -> int:
    for r_off, row_vals in enumerate(rows):
        for c_off, value in enumerate(row_vals):
            cell = ws.cell(row=start_row + r_off, column=1 + c_off, value=value)
            if wrap:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    return start_row + len(rows)


def set_widths(ws: Worksheet, widths: list[int]) -> None:
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def style_status_cell(cell) -> None:
    value = str(cell.value or "")
    if value.startswith("Included"):
        cell.fill = fill(INCLUDED_FILL)
        cell.font = Font(bold=True, color=INCLUDED_TXT)
    elif value.startswith("Excluded"):
        cell.fill = fill(EXCLUDED_FILL)
        cell.font = Font(bold=True, color=EXCLUDED_TXT)
    elif value.startswith("Flag"):
        cell.fill = fill(FLAG_FILL)
        cell.font = Font(bold=True, color=FLAG_TXT)


def project_details_rows(data: dict[str, Any], profile: dict[str, Any] | None) -> list[list[str]]:
    if data.get("project_details"):
        return [list(pair) for pair in data["project_details"]]
    rows: list[list[str]] = []
    for label, key in (
        ("Project", "project_name"),
        ("Address", "address"),
        ("GC / Contact", "gc_contact"),
        ("Bid Due", "bid_due"),
        ("Prepared", "prepared_date"),
    ):
        if data.get(key):
            rows.append([label, str(data[key])])
    if profile:
        rows.append(["Scope Profile", f"{profile['profile_id']} - {profile['trade_name']}"])
        rows.append(["Scope Rule", profile.get("scope_rule", "")])
    if data.get("latest_revision_basis"):
        basis = data["latest_revision_basis"]
        rows.append(["Latest Basis", " ".join(basis) if isinstance(basis, list) else str(basis)])
    return rows


def locate_takeoff_json(project: Path, config: Path | None) -> Path:
    if config:
        return config
    takeoff = project / "bid-package-working" / "takeoff"
    candidates = [
        path
        for path in sorted(takeoff.glob("*.json"))
        if not path.name.startswith("scope-pages-sources") and not path.name.startswith("_")
    ]
    if len(candidates) != 1:
        names = ", ".join(path.name for path in candidates)
        raise FileNotFoundError(
            f"Expected one takeoff JSON in {takeoff}; found {len(candidates)} ({names})."
        )
    return candidates[0]


def check_scope(data: dict[str, Any], profile: dict[str, Any] | None) -> tuple[list[str], list[str]]:
    errors: list[str] = []
    warnings: list[str] = []
    exclude_terms = [term.lower() for term in (profile or {}).get("exclude_terms", [])]
    review_terms = [term.lower() for term in (profile or {}).get("review_terms", [])]
    for idx, row in enumerate(data.get("bom", []), start=1):
        text = " ".join(str(row.get(key, "")) for key in ("section", "item", "description")).lower()
        status = str(row.get("status", ""))
        qty = row.get("qty", 0) or 0
        if status.startswith("Included"):
            for term in exclude_terms:
                if term and term in text:
                    errors.append(
                        f"BOM row {idx}: `{term}` is excluded by profile but marked Included."
                    )
            for term in review_terms:
                if term and term in text:
                    warnings.append(
                        f"BOM row {idx}: `{term}` is review-only by profile but marked Included."
                    )
        if status.startswith("Excluded") and isinstance(qty, (int, float)) and qty:
            warnings.append(f"BOM row {idx}: excluded row has qty {qty}; consider setting qty to 0.")
    return errors, warnings


def build_workbook(
    project: Path,
    *,
    profile: dict[str, Any] | None = None,
    config: Path | None = None,
    out: Path | None = None,
) -> Path:
    if not OPENPYXL_AVAILABLE:
        raise RuntimeError("openpyxl is required to build the workbook. Install with: pip install openpyxl")
    project = project.resolve()
    config_path = locate_takeoff_json(project, config)
    data = read_json(config_path)
    if "bom" not in data:
        raise ValueError(f"No bom[] section found in {config_path}")

    errors, warnings = check_scope(data, profile)
    if errors:
        raise ValueError("Scope violations:\n" + "\n".join(f"- {error}" for error in errors))

    work = project / "bid-package-working"
    out_path = (out or work / "01-Takeoff-Worksheet-REV1.xlsx").resolve()
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    summary = wb.create_sheet("Summary")
    title_block(
        summary,
        f"{data.get('project_name', project.name)} - Takeoff & BOM",
        data.get("scope_rule") or (profile or {}).get("scope_rule", ""),
        4,
    )
    section_header(summary, 4, "Project / Revision Basis", 4)
    row = write_rows(summary, 5, project_details_rows(data, profile))
    for r in range(5, row):
        summary.cell(row=r, column=1).font = Font(bold=True)

    row += 1
    scope_summary = data.get("scope_summary", [])
    if scope_summary:
        section_header(summary, row, "Scope Summary", 4)
        row += 1
        hdr = row
        row = write_rows(
            summary,
            row,
            [["Status", "Scope", "Qty / Size", "Notes"]]
            + [
                [
                    item.get("status", ""),
                    item.get("scope", ""),
                    item.get("quantity", ""),
                    item.get("notes", ""),
                ]
                for item in scope_summary
            ],
        )
        style_header_row(summary, hdr, 4)
        for r in range(hdr + 1, row):
            style_status_cell(summary.cell(row=r, column=1))
    set_widths(summary, [24, 42, 24, 68])
    summary.freeze_panes = "A5"

    bom_ws = wb.create_sheet("BOM")
    title_block(bom_ws, f"{data.get('project_name', project.name)} - BOM", "Yellow columns are supplier/pricing inputs.", 12)
    headers = [
        "Section",
        "Item",
        "Description",
        "Qty",
        "UOM",
        "Status",
        "Source",
        "Supplier",
        "Material $/Unit",
        "Labor $/Unit",
        "Lead Time",
        "Quote Notes",
        "Line Total",
    ]
    write_rows(bom_ws, 4, [headers])
    style_header_row(bom_ws, 4, len(headers))
    start = 5
    for idx, item in enumerate(data.get("bom", []), start=start):
        values = [
            item.get("section", ""),
            item.get("item", ""),
            item.get("description", ""),
            item.get("qty", 0),
            item.get("uom", ""),
            item.get("status", "Included"),
            item.get("source", ""),
            item.get("supplier", ""),
            item.get("material_unit_cost", ""),
            item.get("labor_unit_cost", ""),
            item.get("lead_time", ""),
            item.get("quote_notes", ""),
            f'=IF(LEFT(F{idx},8)="Included",D{idx}*(N(I{idx})+N(J{idx})),0)',
        ]
        write_rows(bom_ws, idx, [values])
        style_status_cell(bom_ws.cell(row=idx, column=6))
        for col in (8, 9, 10, 11, 12):
            bom_ws.cell(row=idx, column=col).fill = fill(INPUT_YELLOW)
    total_row = start + len(data.get("bom", [])) + 1
    bom_ws.cell(row=total_row, column=12, value="Base Bid Total").font = Font(bold=True)
    bom_ws.cell(row=total_row, column=13, value=f"=SUM(M{start}:M{total_row - 2})").font = Font(bold=True)
    set_widths(bom_ws, [18, 32, 48, 12, 10, 18, 28, 22, 16, 16, 18, 36, 16])
    bom_ws.freeze_panes = "A5"

    specs = wb.create_sheet("Scope & Specs")
    title_block(specs, f"{data.get('project_name', project.name)} - Scope & Specs", "", 4)
    spec_rows = [
        [item.get("category", ""), item.get("requirement", ""), item.get("source", ""), item.get("notes", "")]
        for item in data.get("scope_specs", [])
    ]
    write_rows(specs, 4, [["Category", "Requirement", "Source", "Notes"]] + spec_rows)
    style_header_row(specs, 4, 4)
    set_widths(specs, [24, 72, 30, 46])

    refs = wb.create_sheet("Refs & RFIs")
    title_block(refs, f"{data.get('project_name', project.name)} - Refs & RFIs", "", 5)
    ref_rows = [
        [
            item.get("source", ""),
            item.get("pdf_page", ""),
            item.get("sheet", ""),
            item.get("finding", ""),
            item.get("action", ""),
        ]
        for item in data.get("page_references", [])
    ]
    write_rows(refs, 4, [["Source", "PDF Page", "Sheet", "Finding", "Action"]] + ref_rows)
    style_header_row(refs, 4, 5)
    rfi_start = 6 + len(ref_rows)
    section_header(refs, rfi_start, "Open RFIs / Clarifications", 5)
    write_rows(refs, rfi_start + 1, [[note] for note in data.get("rfi_notes", [])])
    set_widths(refs, [36, 12, 18, 64, 46])

    alerts = wb.create_sheet("Alerts")
    title_block(alerts, f"{data.get('project_name', project.name)} - Alerts", "", 3)
    rows = [[item, "Project alert", ""] for item in data.get("alerts", [])]
    rows += [[item, "Profile warning", ""] for item in warnings]
    rows += [[name, "Standard watch list", note] for name, note in WATCH_LIST]
    write_rows(alerts, 4, [["Alert", "Type", "Notes"]] + rows)
    style_header_row(alerts, 4, 3)
    set_widths(alerts, [44, 24, 72])

    sources = wb.create_sheet("Sources")
    title_block(sources, f"{data.get('project_name', project.name)} - Sources", "", 3)
    source_rows = [
        [item.get("file", ""), item.get("pages", ""), item.get("role", "")]
        for item in data.get("source_files", [])
    ]
    write_rows(sources, 4, [["File", "Pages", "Role"]] + source_rows)
    style_header_row(sources, 4, 3)
    set_widths(sources, [70, 12, 50])

    wb.save(out_path)
    return out_path
