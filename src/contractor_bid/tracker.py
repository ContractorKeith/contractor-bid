from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Any

from .util import now_iso, read_json, slugify, today_iso, write_json


# Workspace-level bid tracker. The JSON file is the source of truth; the xlsx is a
# readable, regenerated view. Both live at the workspace root so the tracker spans
# every bid project, not a single one.
TRACKER_JSON = ".contractor-bid/bid-tracker.json"
TRACKER_XLSX = "Bid-Tracker.xlsx"

# Suggested progress stages. Free text is allowed, but these drive cell coloring and
# give the agent/estimator a consistent vocabulary.
PROGRESS_STAGES = [
    "New",
    "Triage",
    "Takeoff",
    "Pricing",
    "Submitted",
    "Won",
    "Lost",
    "No-Bid",
]

OUTCOMES = ["completed", "won", "lost", "no-bid"]

# Reused palette (kept in sync with workbook.py on purpose).
NAVY = "1F4E78"
TEAL = "173F4F"
TEAL_LIGHT = "EAF3F6"
GREEN_FILL = "E2EFDA"
GREEN_TXT = "215E21"
RED_FILL = "FCE4D6"
RED_TXT = "9C3D0A"
AMBER_FILL = "FFF2CC"
AMBER_TXT = "7F6000"

ACTIVE_HEADERS = ["Project", "Location", "Due", "Progress", "Next Action", "Client / GC", "Updated"]
ARCHIVE_HEADERS = ["Project", "Location", "Due", "Outcome", "Client / GC", "Closed"]


def tracker_paths(root: Path) -> tuple[Path, Path]:
    return root / TRACKER_JSON, root / TRACKER_XLSX


def load_tracker(root: Path) -> dict[str, Any]:
    json_path, _ = tracker_paths(root)
    if json_path.exists():
        data = read_json(json_path)
        data.setdefault("bids", [])
        return data
    return {"schema_version": 1, "created": today_iso(), "bids": []}


def save_tracker(root: Path, data: dict[str, Any]) -> None:
    json_path, _ = tracker_paths(root)
    write_json(json_path, data)


def find_entry(data: dict[str, Any], key: str) -> dict[str, Any] | None:
    """Match a bid by id (slug) or case-insensitive project name."""
    key_slug = slugify(key)
    for entry in data.get("bids", []):
        if entry.get("id") == key or entry.get("id") == key_slug:
            return entry
        if str(entry.get("project", "")).strip().lower() == key.strip().lower():
            return entry
    return None


def _load_project_meta(project_path: Path) -> dict[str, Any]:
    """Pull tracker fields from a bid project's project.json when present."""
    meta: dict[str, Any] = {}
    project_json = project_path / "project.json"
    if not project_json.exists():
        return meta
    try:
        raw = read_json(project_json)
    except Exception:
        return meta
    meta["project"] = raw.get("project_name") or project_path.name
    meta["id"] = raw.get("project_slug") or slugify(project_path.name)
    meta["location"] = raw.get("address", "")
    meta["due_date"] = raw.get("bid_due", "")
    meta["client_gc"] = raw.get("gc", "")
    meta["profile"] = raw.get("scope_profile", "")
    return meta


def add_or_update(
    root: Path,
    *,
    project_path: Path | None = None,
    id: str | None = None,
    name: str | None = None,
    location: str | None = None,
    due_date: str | None = None,
    progress: str | None = None,
    next_action: str | None = None,
    client_gc: str | None = None,
    profile: str | None = None,
    note: str | None = None,
) -> tuple[dict[str, Any], bool, list[str]]:
    """Add a new bid or update an existing one. Returns (entry, created, changed_fields)."""
    data = load_tracker(root)

    meta: dict[str, Any] = {}
    rel_path = ""
    if project_path is not None:
        project_path = project_path.resolve()
        meta = _load_project_meta(project_path)
        try:
            rel_path = str(project_path.relative_to(root))
        except ValueError:
            rel_path = str(project_path)

    entry_id = id or meta.get("id") or (slugify(name) if name else None)
    display_name = name or meta.get("project")
    if not entry_id and not display_name:
        raise ValueError("Provide a project path, --id, or --name to identify the bid.")
    if not entry_id:
        entry_id = slugify(display_name)

    # Explicit flags win over project.json; project.json wins over nothing.
    incoming = {
        "project": display_name or meta.get("project") or entry_id,
        "location": location if location is not None else meta.get("location"),
        "due_date": due_date if due_date is not None else meta.get("due_date"),
        "client_gc": client_gc if client_gc is not None else meta.get("client_gc"),
        "profile": profile if profile is not None else meta.get("profile"),
        "progress": progress,
        "next_action": next_action,
    }
    if rel_path:
        incoming["project_path"] = rel_path

    entry = find_entry(data, entry_id) or (find_entry(data, display_name) if display_name else None)
    created = entry is None
    changed: list[str] = []

    if created:
        entry = {
            "id": entry_id,
            "project": incoming["project"],
            "location": incoming.get("location") or "",
            "due_date": incoming.get("due_date") or "",
            "progress": progress or "New",
            "next_action": next_action or "",
            "client_gc": incoming.get("client_gc") or "",
            "profile": incoming.get("profile") or "",
            "project_path": incoming.get("project_path", rel_path),
            "status": "active",
            "outcome": "",
            "notes": [],
            "created": now_iso(),
            "updated": now_iso(),
            "closed": "",
        }
        data["bids"].append(entry)
        changed = ["created"]
    else:
        for field_name in (
            "project",
            "location",
            "due_date",
            "progress",
            "next_action",
            "client_gc",
            "profile",
            "project_path",
        ):
            value = incoming.get(field_name)
            if value is not None and value != "" and value != entry.get(field_name):
                entry[field_name] = value
                changed.append(field_name)
        if changed:
            entry["updated"] = now_iso()

    if note:
        entry.setdefault("notes", []).append({"at": now_iso(), "note": note})
        if "note" not in changed:
            changed.append("note")
        entry["updated"] = now_iso()

    save_tracker(root, data)
    return entry, created, changed


def move_entry(root: Path, key: str, *, outcome: str = "completed") -> dict[str, Any]:
    """Move an active bid to the archived/completed sheet."""
    data = load_tracker(root)
    entry = find_entry(data, key)
    if entry is None:
        raise ValueError(f"No bid found matching: {key}")
    entry["status"] = "archived"
    entry["outcome"] = outcome
    entry["closed"] = today_iso()
    entry["updated"] = now_iso()
    save_tracker(root, data)
    return entry


def reopen_entry(root: Path, key: str) -> dict[str, Any]:
    """Move an archived bid back to active."""
    data = load_tracker(root)
    entry = find_entry(data, key)
    if entry is None:
        raise ValueError(f"No bid found matching: {key}")
    entry["status"] = "active"
    entry["outcome"] = ""
    entry["closed"] = ""
    entry["updated"] = now_iso()
    save_tracker(root, data)
    return entry


def _parse_due(value: str) -> date | None:
    text = str(value or "").strip()
    if not text:
        return None
    # Accept "2026-07-01" or "2026-07-01 14:00".
    head = text.split()[0]
    try:
        return date.fromisoformat(head)
    except ValueError:
        return None


def due_urgency(value: str, today: date | None = None) -> str:
    """Return '', 'soon', or 'past' for cell coloring."""
    due = _parse_due(value)
    if due is None:
        return ""
    today = today or date.today()
    days = (due - today).days
    if days < 0:
        return "past"
    if days <= 2:
        return "soon"
    return ""


def active_bids(data: dict[str, Any]) -> list[dict[str, Any]]:
    return [b for b in data.get("bids", []) if b.get("status", "active") != "archived"]


def archived_bids(data: dict[str, Any]) -> list[dict[str, Any]]:
    return [b for b in data.get("bids", []) if b.get("status") == "archived"]


def render_tracker(root: Path, *, today: date | None = None) -> Path | None:
    """Regenerate Bid-Tracker.xlsx from the JSON. Best effort: returns None if openpyxl is missing."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None

    _, xlsx_path = tracker_paths(root)
    data = load_tracker(root)
    today = today or date.today()

    def fill(color: str) -> "PatternFill":
        return PatternFill(start_color=color, end_color=color, fill_type="solid")

    def style_progress(cell, value: str) -> None:
        low = str(value or "").strip().lower()
        if low in {"won", "submitted", "completed"}:
            cell.fill, cell.font = fill(GREEN_FILL), Font(bold=True, color=GREEN_TXT)
        elif low in {"lost", "no-bid"}:
            cell.fill, cell.font = fill(RED_FILL), Font(bold=True, color=RED_TXT)
        elif low in {"pricing", "takeoff", "triage"}:
            cell.fill, cell.font = fill(AMBER_FILL), Font(bold=True, color=AMBER_TXT)

    def style_due(cell, value: str) -> None:
        urgency = due_urgency(value, today)
        if urgency == "past":
            cell.fill, cell.font = fill(RED_FILL), Font(bold=True, color=RED_TXT)
        elif urgency == "soon":
            cell.fill, cell.font = fill(AMBER_FILL), Font(bold=True, color=AMBER_TXT)

    wb = Workbook()
    wb.remove(wb.active)

    def header_row(ws, headers: list[str]) -> None:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        title = ws.cell(row=1, column=1, value=ws.title)
        title.fill, title.font = fill(TEAL), Font(bold=True, color="FFFFFF", size=14)
        title.alignment = Alignment(horizontal="center", vertical="center")
        sub = ws.cell(
            row=2,
            column=1,
            value=f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')} by contractor-bid · source: {TRACKER_JSON}",
        )
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
        sub.fill, sub.font = fill(TEAL_LIGHT), Font(italic=True, color=TEAL, size=9)
        for col, name in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col, value=name)
            cell.fill = fill(NAVY)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # --- Active Bids sheet ---
    active = wb.create_sheet("Active Bids")
    header_row(active, ACTIVE_HEADERS)
    row = 5
    for entry in active_bids(data):
        values = [
            entry.get("project", ""),
            entry.get("location", ""),
            entry.get("due_date", ""),
            entry.get("progress", ""),
            entry.get("next_action", ""),
            entry.get("client_gc", ""),
            (entry.get("updated", "") or "")[:10],
        ]
        for col, value in enumerate(values, start=1):
            cell = active.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        style_due(active.cell(row=row, column=3), entry.get("due_date", ""))
        style_progress(active.cell(row=row, column=4), entry.get("progress", ""))
        row += 1
    if row == 5:
        active.cell(row=5, column=1, value="No active bids yet.").font = Font(italic=True)
    for idx, width in enumerate([30, 26, 13, 14, 44, 30, 12], start=1):
        active.column_dimensions[get_column_letter(idx)].width = width
    active.freeze_panes = "A5"

    # --- Archived / Completed sheet ---
    # Excel sheet titles cannot contain / \ ? * [ ] : — use "&" for the label.
    archive = wb.create_sheet("Archived & Completed")
    header_row(archive, ARCHIVE_HEADERS)
    row = 5
    for entry in archived_bids(data):
        values = [
            entry.get("project", ""),
            entry.get("location", ""),
            entry.get("due_date", ""),
            entry.get("outcome", ""),
            entry.get("client_gc", ""),
            entry.get("closed", ""),
        ]
        for col, value in enumerate(values, start=1):
            cell = archive.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        style_progress(archive.cell(row=row, column=4), entry.get("outcome", ""))
        row += 1
    if row == 5:
        archive.cell(row=5, column=1, value="No archived bids yet.").font = Font(italic=True)
    for idx, width in enumerate([30, 26, 13, 14, 30, 12], start=1):
        archive.column_dimensions[get_column_letter(idx)].width = width
    archive.freeze_panes = "A5"

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)
    return xlsx_path


def change_summary(entry: dict[str, Any], created: bool, changed: list[str]) -> str:
    """One-line human summary of what an add/update did (for the agent to echo back)."""
    name = entry.get("project", entry.get("id", "bid"))
    if created:
        return (
            f"Added bid '{name}' to tracker (Active): "
            f"progress={entry.get('progress', '')}, due={entry.get('due_date') or 'TBD'}."
        )
    if not changed:
        return f"No changes for bid '{name}'."
    parts = [f"{field}={entry.get(field, '')}" for field in changed if field not in {"note"}]
    if "note" in changed:
        parts.append("added note")
    return f"Updated bid '{name}': " + ", ".join(parts) + "."
